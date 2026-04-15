"""Reports router.

Exposes per-deal and portfolio-level report exports:

- ``GET /api/reports/deal/{deal_id}/pdf`` — PDF summary of a single deal.
- ``GET /api/reports/portfolio/excel`` — Excel workbook of all investments.
- ``GET /api/reports/portfolio/quarterly/pdf`` — Quarterly PDF performance report.
- ``GET /api/reports/export/json`` — Full JSON dump (deals + investments).
- ``GET /api/reports/export/csv`` — CSV dump (two sheets joined into one zip).

PDF generation uses reportlab, Excel uses openpyxl, JSON/CSV use stdlib.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models import Deal, Developer, Investment
from app.services.portfolio_analytics import portfolio_analytics

router = APIRouter()


# ===== Helpers =====

def _fmt_money(v) -> str:
    if v is None:
        return "—"
    try:
        return f"${float(v):,.0f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):.2f}%"
    except (TypeError, ValueError):
        return str(v)


def _pdf_style():
    """Shared reportlab styles for consistent branding across reports."""
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="BrandTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.HexColor("#1a1a2e"),
        spaceAfter=8,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="SectionHeading",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#4361ee"),
        spaceBefore=12,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="Small",
        parent=styles["BodyText"],
        fontSize=9,
        textColor=colors.HexColor("#64748b"),
    ))
    return styles


def _metric_table(pairs: list[tuple[str, str]]):
    """Build a 2-column metric table from (label, value) pairs."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors

    t = Table(pairs, colWidths=[170, 300])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#1a1a2e")),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
    ]))
    return t


# ===== Deal PDF =====

@router.get("/deal/{deal_id}/pdf")
async def deal_pdf_report(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Generate a PDF summary report for a single deal."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak

    result = await db.execute(
        select(Deal)
        .options(selectinload(Deal.developer), selectinload(Deal.documents))
        .where(Deal.id == deal_id)
    )
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    metrics = deal.metrics or {}
    scores = deal.scores or {}
    ds = metrics.get("deal_structure", {}) or {}
    tr = metrics.get("target_returns", {}) or {}
    pd_ = metrics.get("project_details", {}) or {}
    fp = metrics.get("financial_projections", {}) or {}
    ml = metrics.get("market_location", {}) or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=48, rightMargin=48, topMargin=48, bottomMargin=48,
        title=f"{deal.project_name} Deal Report",
    )
    styles = _pdf_style()
    story = []

    # Header
    story.append(Paragraph(deal.project_name, styles["BrandTitle"]))
    subtitle_parts = [
        deal.developer.name if deal.developer else None,
        deal.city or ml.get("city"),
        deal.state or ml.get("state"),
        (deal.property_type or "").title(),
    ]
    subtitle = " · ".join(p for p in subtitle_parts if p)
    if subtitle:
        story.append(Paragraph(subtitle, styles["Small"]))
    story.append(Paragraph(
        f"Generated {date.today().isoformat()} · Status: {deal.status or 'reviewing'}",
        styles["Small"],
    ))
    story.append(Spacer(1, 12))

    # Scores
    story.append(Paragraph("Investment Scores", styles["SectionHeading"]))
    if scores:
        rows = [("Overall", f"{scores.get('overall', '—')} / 10")]
        for cat in ["returns", "market", "structure", "risk", "financials", "underwriting", "sponsor"]:
            c = scores.get(cat) or {}
            if c.get("score") is not None:
                rows.append((cat.title(), f"{c.get('score')} / 10  ({c.get('weight', 0)}% weight)"))
        story.append(_metric_table(rows))
        # Score notes
        for cat in ["returns", "market", "structure", "risk", "financials", "underwriting", "sponsor"]:
            c = scores.get(cat) or {}
            if c.get("notes"):
                story.append(Paragraph(f"<b>{cat.title()}:</b> {c['notes']}", styles["Small"]))
    else:
        story.append(Paragraph("No scores computed yet.", styles["Small"]))

    # Deal structure
    story.append(Paragraph("Deal Structure", styles["SectionHeading"]))
    story.append(_metric_table([
        ("Investment Class", ds.get("investment_class") or "—"),
        ("Minimum Investment", _fmt_money(ds.get("minimum_investment"))),
        ("Total Project Cost", _fmt_money(ds.get("total_project_cost"))),
        ("Total Equity", _fmt_money(ds.get("total_equity_required"))),
        ("Debt Amount", _fmt_money(ds.get("debt_amount"))),
        ("LTV", _fmt_pct(ds.get("ltv"))),
        ("Interest Rate", _fmt_pct(ds.get("interest_rate"))),
        ("Hold Period", f"{ds['hold_period_years']} years" if ds.get("hold_period_years") else "—"),
        ("Preferred Return", _fmt_pct(ds.get("preferred_return"))),
        ("GP Co-Invest", str(ds.get("gp_coinvest") or "—")),
        ("Asset Mgmt Fee", _fmt_pct(ds.get("fees_asset_mgmt"))),
    ]))

    # Target returns
    story.append(Paragraph("Target Returns", styles["SectionHeading"]))
    story.append(_metric_table([
        ("Target IRR", _fmt_pct(tr.get("target_irr"))),
        ("Target Equity Multiple", f"{tr['target_equity_multiple']}x" if tr.get("target_equity_multiple") else "—"),
        ("Cash-on-Cash", _fmt_pct(tr.get("target_cash_on_cash"))),
        ("Avg Annual Return", _fmt_pct(tr.get("target_avg_annual_return"))),
        ("Projected Profit", _fmt_money(tr.get("projected_profit"))),
    ]))

    # Project details
    story.append(Paragraph("Project Details", styles["SectionHeading"]))
    story.append(_metric_table([
        ("Unit Count", str(pd_.get("unit_count") or "—")),
        ("Total SqFt", f"{pd_['total_sqft']:,}" if pd_.get("total_sqft") else "—"),
        ("Price / Unit", _fmt_money(pd_.get("price_per_unit"))),
        ("Price / SqFt", _fmt_money(pd_.get("price_per_sqft"))),
        ("Construction Type", pd_.get("construction_type") or "—"),
        ("Entitlement Status", pd_.get("entitlement_status") or "—"),
    ]))

    # Financial projections
    story.append(Paragraph("Financial Projections", styles["SectionHeading"]))
    story.append(_metric_table([
        ("Stabilized NOI", _fmt_money(fp.get("stabilized_noi"))),
        ("Entry Cap Rate", _fmt_pct(fp.get("entry_cap_rate"))),
        ("Exit Cap Rate", _fmt_pct(fp.get("exit_cap_rate"))),
        ("Avg Rent / Unit", _fmt_money(fp.get("avg_rent_per_unit"))),
        ("Rent Growth", _fmt_pct(fp.get("rent_growth_assumption"))),
        ("Occupancy", _fmt_pct(fp.get("occupancy_assumption"))),
    ]))

    # Notes
    if deal.notes:
        story.append(PageBreak())
        story.append(Paragraph("Notes", styles["SectionHeading"]))
        story.append(Paragraph(deal.notes.replace("\n", "<br/>"), styles["BodyText"]))

    # Documents
    if deal.documents:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Source Documents", styles["SectionHeading"]))
        for d in deal.documents:
            line = f"• {d.filename}  ({d.doc_type}, {d.page_count} pages)"
            story.append(Paragraph(line, styles["Small"]))

    doc.build(story)
    buf.seek(0)

    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (deal.project_name or f"deal_{deal_id}"))
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe}_report.pdf"'},
    )


# ===== Portfolio Excel =====

@router.get("/portfolio/excel")
async def portfolio_excel_report(db: AsyncSession = Depends(get_db)):
    """Excel workbook with investments, distributions, and performance sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    result = await db.execute(
        select(Investment).options(selectinload(Investment.distributions))
    )
    investments = result.scalars().all()

    analytics = portfolio_analytics(investments)
    perf_by_id = {p["investment_id"]: p for p in analytics["per_investment"]}

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    def _write_header(ws, headers, widths=None):
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            if widths:
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = widths[idx - 1]

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    s = analytics["summary"]
    ws["A1"] = "Portfolio Summary"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Total Investments", s["investment_count"]),
        ("Total Invested", s["total_invested"]),
        ("Total Distributions", s["total_distributions"]),
        ("Total Exit Proceeds", s["total_exit_proceeds"]),
        ("Total Returned", s["total_returned"]),
        ("Net Profit", s["net_profit"]),
        ("Overall Multiple", s["overall_multiple"]),
        ("Portfolio IRR (%)", s["overall_irr_pct"]),
    ]
    for i, (label, val) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    # --- Sheet 2: Investments ---
    ws = wb.create_sheet("Investments")
    headers = [
        "ID", "Project", "Sponsor", "Class", "Status",
        "Invest Date", "Invested", "Distributions", "Exit Amount", "Total Returned",
        "Net Profit", "Multiple", "DPI", "IRR (%)", "Years Held",
        "Projected IRR (%)", "Projected Multiple", "IRR vs Projected (pp)",
        "Pref Return (%)", "Hold Period (yrs)", "Notes",
    ]
    _write_header(ws, headers, widths=[6, 26, 22, 14, 12, 12, 14, 14, 14, 14, 14, 10, 8, 10, 10, 14, 14, 18, 12, 12, 40])
    for idx, inv in enumerate(investments, 2):
        perf = perf_by_id.get(inv.id, {})
        ws.cell(row=idx, column=1, value=inv.id)
        ws.cell(row=idx, column=2, value=inv.project_name)
        ws.cell(row=idx, column=3, value=inv.sponsor_name)
        ws.cell(row=idx, column=4, value=inv.investment_class)
        ws.cell(row=idx, column=5, value=inv.status)
        ws.cell(row=idx, column=6, value=inv.investment_date)
        ws.cell(row=idx, column=7, value=inv.amount_invested)
        ws.cell(row=idx, column=8, value=perf.get("total_distributions"))
        ws.cell(row=idx, column=9, value=perf.get("exit_amount"))
        ws.cell(row=idx, column=10, value=perf.get("total_returned"))
        ws.cell(row=idx, column=11, value=perf.get("net_profit"))
        ws.cell(row=idx, column=12, value=perf.get("multiple"))
        ws.cell(row=idx, column=13, value=perf.get("dpi"))
        ws.cell(row=idx, column=14, value=perf.get("irr"))
        ws.cell(row=idx, column=15, value=perf.get("years_held"))
        ws.cell(row=idx, column=16, value=perf.get("projected_irr"))
        ws.cell(row=idx, column=17, value=perf.get("projected_multiple"))
        ws.cell(row=idx, column=18, value=perf.get("irr_vs_projected"))
        ws.cell(row=idx, column=19, value=inv.preferred_return)
        ws.cell(row=idx, column=20, value=inv.hold_period_years)
        ws.cell(row=idx, column=21, value=inv.notes)

    # --- Sheet 3: Distributions ---
    ws = wb.create_sheet("Distributions")
    _write_header(
        ws,
        ["Investment ID", "Project", "Date", "Amount", "Type", "Period", "Notes"],
        widths=[12, 26, 12, 14, 18, 14, 40],
    )
    r = 2
    for inv in investments:
        for d in sorted(inv.distributions or [], key=lambda x: x.date or date.min):
            ws.cell(row=r, column=1, value=inv.id)
            ws.cell(row=r, column=2, value=inv.project_name)
            ws.cell(row=r, column=3, value=d.date)
            ws.cell(row=r, column=4, value=d.amount)
            ws.cell(row=r, column=5, value=d.dist_type)
            ws.cell(row=r, column=6, value=d.period)
            ws.cell(row=r, column=7, value=d.notes)
            r += 1

    # --- Sheet 4: By Sponsor ---
    ws = wb.create_sheet("By Sponsor")
    _write_header(ws, ["Sponsor", "# Investments", "Invested", "Returned", "Multiple", "Share %"],
                  widths=[28, 14, 16, 16, 10, 10])
    for i, row in enumerate(analytics["by_sponsor"], 2):
        ws.cell(row=i, column=1, value=row["name"])
        ws.cell(row=i, column=2, value=row["count"])
        ws.cell(row=i, column=3, value=row["invested"])
        ws.cell(row=i, column=4, value=row["returned"])
        ws.cell(row=i, column=5, value=row["multiple"])
        ws.cell(row=i, column=6, value=row["share_pct"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="portfolio_{date.today().isoformat()}.xlsx"'},
    )


# ===== Quarterly PDF =====

def _quarter_key(d: date) -> str:
    return f"{d.year}-Q{(d.month - 1) // 3 + 1}"


@router.get("/portfolio/quarterly/pdf")
async def portfolio_quarterly_pdf(db: AsyncSession = Depends(get_db)):
    """Quarterly performance PDF: distributions bucketed by quarter."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    result = await db.execute(
        select(Investment).options(selectinload(Investment.distributions))
    )
    investments = result.scalars().all()
    analytics = portfolio_analytics(investments)
    s = analytics["summary"]

    # Bucket distributions & exits by quarter
    quarters: dict[str, dict] = {}
    for inv in investments:
        for d in inv.distributions or []:
            if d.date and d.amount:
                q = _quarter_key(d.date)
                b = quarters.setdefault(q, {"distributions": 0.0, "exit": 0.0, "new_invested": 0.0})
                b["distributions"] += float(d.amount)
        if inv.exit_date and inv.exit_amount:
            q = _quarter_key(inv.exit_date)
            b = quarters.setdefault(q, {"distributions": 0.0, "exit": 0.0, "new_invested": 0.0})
            b["exit"] += float(inv.exit_amount)
        if inv.investment_date and inv.amount_invested:
            q = _quarter_key(inv.investment_date)
            b = quarters.setdefault(q, {"distributions": 0.0, "exit": 0.0, "new_invested": 0.0})
            b["new_invested"] += float(inv.amount_invested)

    sorted_quarters = sorted(quarters.items())

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=48, rightMargin=48, topMargin=48, bottomMargin=48)
    styles = _pdf_style()
    story = []

    story.append(Paragraph("Portfolio Performance Report", styles["BrandTitle"]))
    story.append(Paragraph(f"Generated {date.today().isoformat()}", styles["Small"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Summary", styles["SectionHeading"]))
    story.append(_metric_table([
        ("Total Investments", str(s["investment_count"])),
        ("Total Invested", _fmt_money(s["total_invested"])),
        ("Total Distributions", _fmt_money(s["total_distributions"])),
        ("Total Exit Proceeds", _fmt_money(s["total_exit_proceeds"])),
        ("Total Returned", _fmt_money(s["total_returned"])),
        ("Net Profit", _fmt_money(s["net_profit"])),
        ("Overall Multiple", f"{s['overall_multiple']}x"),
        ("Portfolio IRR", _fmt_pct(s["overall_irr_pct"])),
    ]))

    story.append(Paragraph("Quarterly Cash Flows", styles["SectionHeading"]))
    if sorted_quarters:
        table_data = [["Quarter", "New Invested", "Distributions", "Exit Proceeds", "Net Flow"]]
        for q, b in sorted_quarters:
            net = b["distributions"] + b["exit"] - b["new_invested"]
            table_data.append([
                q,
                _fmt_money(b["new_invested"]) if b["new_invested"] else "—",
                _fmt_money(b["distributions"]) if b["distributions"] else "—",
                _fmt_money(b["exit"]) if b["exit"] else "—",
                _fmt_money(net),
            ])
        tbl = Table(table_data, colWidths=[70, 90, 95, 90, 90])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("No dated cash flows yet.", styles["Small"]))

    # Top performers
    story.append(Paragraph("Top Performers (by IRR)", styles["SectionHeading"]))
    top = analytics["top_performers"]
    if top:
        by_id = {inv.id: inv for inv in investments}
        data = [["Investment", "Sponsor", "IRR", "Multiple", "Years"]]
        for p in top[:10]:
            inv = by_id.get(p["investment_id"])
            if not inv:
                continue
            data.append([
                inv.project_name or "—",
                inv.sponsor_name or "—",
                f"{p['irr']:.2f}%",
                f"{p['multiple']}x",
                f"{p['years_held'] or '—'}",
            ])
        tbl = Table(data, colWidths=[160, 140, 60, 60, 50])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph("Need at least 2 dated cash flows per investment to compute IRR.", styles["Small"]))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="quarterly_{date.today().isoformat()}.pdf"'},
    )


# ===== Bulk JSON / CSV =====

def _json_default(obj):
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    raise TypeError(f"Not serializable: {type(obj)}")


@router.get("/export/json")
async def export_all_json(db: AsyncSession = Depends(get_db)):
    """Bulk JSON dump of developers, deals, investments and distributions."""
    devs_r = await db.execute(select(Developer))
    developers = [{
        "id": d.id, "name": d.name, "contact_name": d.contact_name,
        "contact_email": d.contact_email, "phone": d.phone,
        "track_record": d.track_record, "notes": d.notes,
        "created_at": d.created_at,
    } for d in devs_r.scalars().all()]

    deals_r = await db.execute(select(Deal).options(selectinload(Deal.documents)))
    deals = [{
        "id": d.id, "developer_id": d.developer_id, "project_name": d.project_name,
        "location": d.location, "city": d.city, "state": d.state,
        "property_type": d.property_type, "status": d.status,
        "metrics": d.metrics, "scores": d.scores, "notes": d.notes,
        "created_at": d.created_at,
        "documents": [{
            "id": doc.id, "filename": doc.filename, "doc_type": doc.doc_type,
            "page_count": doc.page_count, "upload_date": doc.upload_date,
        } for doc in d.documents],
    } for d in deals_r.scalars().all()]

    inv_r = await db.execute(select(Investment).options(selectinload(Investment.distributions)))
    investments = [{
        "id": i.id, "deal_id": i.deal_id, "project_name": i.project_name,
        "sponsor_name": i.sponsor_name, "investment_date": i.investment_date,
        "amount_invested": i.amount_invested, "shares": i.shares,
        "investment_class": i.investment_class, "preferred_return": i.preferred_return,
        "projected_irr": i.projected_irr, "projected_equity_multiple": i.projected_equity_multiple,
        "hold_period_years": i.hold_period_years, "status": i.status,
        "exit_date": i.exit_date, "exit_amount": i.exit_amount, "notes": i.notes,
        "created_at": i.created_at,
        "distributions": [{
            "id": d.id, "date": d.date, "amount": d.amount,
            "dist_type": d.dist_type, "period": d.period, "notes": d.notes,
        } for d in i.distributions],
    } for i in inv_r.scalars().all()]

    payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "developers": developers,
        "deals": deals,
        "investments": investments,
    }
    data = json.dumps(payload, default=_json_default, indent=2)
    return StreamingResponse(
        io.BytesIO(data.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="export_{date.today().isoformat()}.json"'},
    )


@router.get("/export/csv")
async def export_all_csv(db: AsyncSession = Depends(get_db)):
    """Bulk CSV export as a zip (one CSV per table)."""
    devs_r = await db.execute(select(Developer))
    developers = devs_r.scalars().all()
    deals_r = await db.execute(select(Deal))
    deals = deals_r.scalars().all()
    inv_r = await db.execute(select(Investment).options(selectinload(Investment.distributions)))
    investments = inv_r.scalars().all()

    def _csv(rows: list[list], header: list[str]) -> bytes:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(header)
        for row in rows:
            w.writerow([("" if v is None else v) for v in row])
        return out.getvalue().encode("utf-8")

    dev_bytes = _csv(
        [[d.id, d.name, d.contact_name, d.contact_email, d.phone, d.track_record, d.notes,
          d.created_at.isoformat() if d.created_at else ""] for d in developers],
        ["id", "name", "contact_name", "contact_email", "phone", "track_record", "notes", "created_at"],
    )
    deal_bytes = _csv(
        [[d.id, d.developer_id, d.project_name, d.location, d.city, d.state, d.property_type,
          d.status, (d.scores or {}).get("overall"), d.notes,
          d.created_at.isoformat() if d.created_at else ""] for d in deals],
        ["id", "developer_id", "project_name", "location", "city", "state", "property_type",
         "status", "overall_score", "notes", "created_at"],
    )
    inv_bytes = _csv(
        [[i.id, i.deal_id, i.project_name, i.sponsor_name,
          i.investment_date.isoformat() if i.investment_date else "",
          i.amount_invested, i.investment_class, i.preferred_return,
          i.projected_irr, i.projected_equity_multiple, i.hold_period_years, i.status,
          i.exit_date.isoformat() if i.exit_date else "",
          i.exit_amount, i.notes] for i in investments],
        ["id", "deal_id", "project_name", "sponsor_name", "investment_date",
         "amount_invested", "investment_class", "preferred_return",
         "projected_irr", "projected_equity_multiple", "hold_period_years", "status",
         "exit_date", "exit_amount", "notes"],
    )
    dist_rows = []
    for i in investments:
        for d in i.distributions or []:
            dist_rows.append([
                d.id, i.id, i.project_name,
                d.date.isoformat() if d.date else "",
                d.amount, d.dist_type, d.period, d.notes,
            ])
    dist_bytes = _csv(
        dist_rows,
        ["id", "investment_id", "investment_project", "date", "amount", "dist_type", "period", "notes"],
    )

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("developers.csv", dev_bytes)
        zf.writestr("deals.csv", deal_bytes)
        zf.writestr("investments.csv", inv_bytes)
        zf.writestr("distributions.csv", dist_bytes)
    zbuf.seek(0)
    return StreamingResponse(
        zbuf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="export_csv_{date.today().isoformat()}.zip"'},
    )

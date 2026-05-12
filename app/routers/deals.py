import os
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from app.rate_limit import limit
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from pydantic import BaseModel, Field, field_validator
from typing import Optional
from app.database import get_db
from app.models import Deal, DealDocument
from app.services.deal_validator import validate_deal_metrics
from app.services.math_checker import run_math_checks
from app.services.market_data import fetch_market_data
from app.services.cashflow_projector import project_cash_flows
from app.services.waterfall_calculator import waterfall_from_deal
from app.services.canonical_metrics import canonical_return_summary
from app.services.data_integrity import quality_summary
from app.services.location_intelligence import build_location_bundle

router = APIRouter()


# Allowed status values for a Deal - closes the "any string accepted" hole.
# Keeps the legacy "reviewing/interested/passed/committed/closed" taxonomy.
DEAL_STATUSES = {"reviewing", "interested", "passed", "committed", "closed"}
PROPERTY_TYPES = {
    "multifamily", "office", "retail", "industrial", "hospitality",
    "mixed-use", "development", "land", "other",
}


class DealCreate(BaseModel):
    developer_id: Optional[int] = Field(None, ge=1)
    project_name: str = Field(..., min_length=1, max_length=255)
    location: Optional[str] = Field("", max_length=500)
    city: Optional[str] = Field("", max_length=120)
    state: Optional[str] = Field("", max_length=64)
    property_type: Optional[str] = "multifamily"
    status: Optional[str] = "reviewing"
    notes: Optional[str] = Field("", max_length=10000)

    @field_validator("status")
    @classmethod
    def _status_valid(cls, v):
        if v is None or v == "":
            return "reviewing"
        if v not in DEAL_STATUSES:
            raise ValueError(f"status must be one of {sorted(DEAL_STATUSES)}")
        return v

    @field_validator("property_type")
    @classmethod
    def _ptype_valid(cls, v):
        if v is None or v == "":
            return "multifamily"
        # Allow unknown types but normalize to lowercase. Unknown types pass
        # through (e.g. "self-storage") so the system stays extensible.
        return v.strip().lower()


class DealUpdate(BaseModel):
    developer_id: Optional[int] = Field(None, ge=1)
    project_name: Optional[str] = Field(None, min_length=1, max_length=255)
    location: Optional[str] = Field(None, max_length=500)
    city: Optional[str] = Field(None, max_length=120)
    state: Optional[str] = Field(None, max_length=64)
    property_type: Optional[str] = None
    status: Optional[str] = None
    metrics: Optional[dict] = None
    scores: Optional[dict] = None
    notes: Optional[str] = Field(None, max_length=10000)

    @field_validator("status")
    @classmethod
    def _status_valid(cls, v):
        if v is None:
            return v
        if v not in DEAL_STATUSES:
            raise ValueError(f"status must be one of {sorted(DEAL_STATUSES)}")
        return v


class CompareRequest(BaseModel):
    # Bound the compare set so the UI can't force the backend to fan out
    # to unbounded deals (and produce a 40-column Excel file).
    deal_ids: list[int] = Field(..., min_length=1, max_length=8)


def _deal_to_dict(deal: Deal, developer_name: str = None) -> dict:
    metrics = deal.metrics or {}
    scores = deal.scores or {}
    return_summary = canonical_return_summary(metrics)
    deal_structure = metrics.get("deal_structure", {}) or {}

    return {
        "id": deal.id,
        "developer_id": deal.developer_id,
        "developer_name": developer_name or "",
        "project_name": deal.project_name,
        "location": deal.location,
        "city": deal.city,
        "state": deal.state,
        "property_type": deal.property_type,
        "status": deal.status,
        "metrics": metrics,
        "scores": scores,
        "overall_score": scores.get("overall", None),
        "target_irr": return_summary.get("target_irr"),
        "target_equity_multiple": return_summary.get("target_equity_multiple"),
        "minimum_investment": deal_structure.get("minimum_investment"),
        "notes": deal.notes,
        "lat": deal.lat,
        "lng": deal.lng,
        "created_at": deal.created_at.isoformat() if deal.created_at else None,
    }


@router.get("")
async def list_deals(
    trash: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """List deals. Trashed (soft-deleted) rows are excluded unless
    `?trash=true` is passed - handy for a future 'Trash' view."""
    q = select(Deal).options(selectinload(Deal.developer)).order_by(Deal.created_at.desc())
    if trash:
        q = q.where(Deal.deleted_at.is_not(None))
    else:
        q = q.where(Deal.deleted_at.is_(None))
    result = await db.execute(q)
    deals = result.scalars().all()
    return [
        _deal_to_dict(d, d.developer.name if d.developer else "")
        for d in deals
    ]


@router.post("")
async def create_deal(data: DealCreate, db: AsyncSession = Depends(get_db)):
    deal = Deal(**data.model_dump())
    db.add(deal)
    await db.commit()
    await db.refresh(deal)
    return {"id": deal.id, "project_name": deal.project_name, "message": "Deal created"}


@router.get("/pipeline/summary")
async def pipeline_summary_endpoint(
    response: Response,
    db: AsyncSession = Depends(get_db),
):
    """Dashboard widgets: total deals, velocity (6mo), win rate (12mo),
    aging deals, capital deployed, average analyst score. Derived from
    the live Deal table on every call - no materialized view yet.

    Cached for 30s. The numbers change only when a deal is created /
    scored / status-changed, which is infrequent on an operator tool.
    """
    from app.services.pipeline_analytics import pipeline_summary

    result = await db.execute(
        select(Deal).options(selectinload(Deal.developer)).where(Deal.deleted_at.is_(None))
    )
    response.headers["Cache-Control"] = "private, max-age=30"
    return pipeline_summary(result.scalars().all())


@router.get("/{deal_id}")
async def get_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Deal)
        .options(selectinload(Deal.developer), selectinload(Deal.documents))
        .where(Deal.id == deal_id)
    )
    deal = result.scalar_one_or_none()
    # Treat soft-deleted rows as gone for GETs; restore via POST /{id}/restore.
    if not deal or deal.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Deal not found")

    data = _deal_to_dict(deal, deal.developer.name if deal.developer else "")
    data["documents"] = [
        {
            "id": doc.id,
            "filename": doc.filename,
            "doc_type": doc.doc_type,
            "page_count": doc.page_count,
            "upload_date": doc.upload_date.isoformat() if doc.upload_date else None,
            "has_text": bool(doc.extracted_text),
            "extraction_quality": {
                "quality_score": (doc.extraction_quality or {}).get("quality_score"),
                "ocr_pages": (doc.extraction_quality or {}).get("ocr_pages", 0),
                "empty_pages": (doc.extraction_quality or {}).get("empty_pages", []),
            }
            if doc.extraction_quality
            else None,
        }
        for doc in deal.documents
    ]
    # Quality summary of the metrics (counts of verified / extracted / conflicting / locked)
    if deal.metrics:
        data["quality"] = quality_summary(deal.metrics)
    return data


@router.put("/{deal_id}")
async def update_deal(deal_id: int, data: DealUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if value is not None:
            setattr(deal, key, value)
    await db.commit()
    await db.refresh(deal)
    return {"message": "Deal updated", "id": deal.id}


@router.delete("/{deal_id}")
async def delete_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Soft-delete. The deal is hidden from list/get endpoints but
    remains in the DB. `POST /{id}/restore` reverses within the undo
    window; `DELETE /{id}/purge` hard-deletes immediately."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal or deal.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Deal not found")
    deal.deleted_at = datetime.now(timezone.utc)
    await db.commit()
    return {"message": "Deal moved to trash", "id": deal_id}


@router.post("/{deal_id}/restore")
async def restore_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    deal.deleted_at = None
    await db.commit()
    return {"message": "Restored", "id": deal_id}


@router.delete("/{deal_id}/purge")
async def purge_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Hard-delete a soft-deleted deal. Only allowed once the row is
    already in the trash - prevents accidental irreversible removal."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    if deal.deleted_at is None:
        raise HTTPException(
            status_code=409,
            detail="Deal must be in the trash before it can be purged. Call DELETE first.",
        )
    await db.delete(deal)
    await db.commit()
    return {"message": "Purged"}


# ===== Documents =====

@router.get("/{deal_id}/documents")
async def list_documents(deal_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(DealDocument).where(DealDocument.deal_id == deal_id).order_by(DealDocument.upload_date.desc())
    )
    docs = result.scalars().all()
    out = []
    for doc in docs:
        q = doc.extraction_quality or {}
        out.append(
            {
                "id": doc.id,
                "filename": doc.filename,
                "doc_type": doc.doc_type,
                "page_count": doc.page_count,
                "upload_date": doc.upload_date.isoformat() if doc.upload_date else None,
                "has_text": bool(doc.extracted_text),
                "extraction_quality": {
                    "quality_score": q.get("quality_score"),
                    "ocr_pages": q.get("ocr_pages", 0),
                    "empty_pages": q.get("empty_pages", []),
                }
                if q
                else None,
            }
        )
    return out


@router.delete("/documents/{doc_id}")
async def delete_document(doc_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DealDocument).where(DealDocument.id == doc_id))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Delete file
    if os.path.exists(doc.file_path):
        os.remove(doc.file_path)

    await db.delete(doc)
    await db.commit()
    return {"message": "Document deleted"}


@router.get("/documents/{doc_id}/text")
async def get_document_text(doc_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DealDocument).where(DealDocument.id == doc_id))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"id": doc.id, "filename": doc.filename, "text": doc.extracted_text}


# ===== Validation and math =====

@router.get("/{deal_id}/validate")
async def validate_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Run validation checks on current metrics and return flags."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if not deal.metrics:
        raise HTTPException(status_code=400, detail="No metrics extracted yet. Run extraction first.")

    flags = validate_deal_metrics(deal.metrics, deal.property_type)

    # Also update stored flags
    metrics = deal.metrics.copy()
    metrics["validation_flags"] = flags
    deal.metrics = metrics
    await db.commit()

    return {"flags": flags, "summary": {
        "red": len([f for f in flags if f["severity"] == "red"]),
        "yellow": len([f for f in flags if f["severity"] == "yellow"]),
        "green": len([f for f in flags if f["severity"] == "green"]),
    }}


@router.get("/{deal_id}/math-check")
async def math_check_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Run deterministic math verification - zero AI, pure arithmetic.
    
    Cross-checks all calculations, internal consistency, and benchmark ranges.
    """
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if not deal.metrics:
        raise HTTPException(status_code=400, detail="No metrics extracted yet.")

    checks = run_math_checks(deal.metrics)
    
    summary = {
        'pass': len([c for c in checks if c['status'] == 'pass']),
        'fail': len([c for c in checks if c['status'] == 'fail']),
        'warn': len([c for c in checks if c['status'] == 'warn']),
        'info': len([c for c in checks if c['status'] == 'info']),
        'total': len(checks),
    }

    return {"checks": checks, "summary": summary}


# ===== Location intelligence =====

@router.get("/{deal_id}/location")
async def get_deal_location(
    deal_id: int,
    radius_m: int = 1600,
    refresh: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """Return cached or freshly-fetched location data for a deal.

    Payload shape (all sources free / unauthenticated by default):
      - lat, lng             - resolved via Nominatim or user-placed
      - display_name         - free-form, good enough for a map attribution
      - radius_m             - currently fetched radius
      - categories           - {apartments|restaurants|grocery|transit|schools|
                                healthcare|parks|employers: [POI...]}
      - fmr                  - HUD Fair Market Rent (if HUD_API_TOKEN set)
      - fetched_at           - unix timestamp, used for staleness UI

    Results are cached in `deal.location_data` for 7 days unless
    `refresh=true` is passed.
    """
    radius_m = max(500, min(8000, int(radius_m)))
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    bundle = await build_location_bundle(deal, radius_m=radius_m, force_refresh=refresh)

    # Persist whatever we learned so the next page load is instant.
    if bundle.get("lat") is not None and bundle.get("lng") is not None:
        deal.lat = bundle["lat"]
        deal.lng = bundle["lng"]
    deal.location_data = bundle
    await db.commit()

    return bundle


class LocationManualIn(BaseModel):
    lat: float = Field(..., ge=-90, le=90)
    lng: float = Field(..., ge=-180, le=180)


@router.post("/{deal_id}/location/manual")
async def set_manual_location(
    deal_id: int,
    data: LocationManualIn,
    db: AsyncSession = Depends(get_db),
):
    """Pin the deal's map position by hand.

    Geocoding can miss when the address is a new development or when the
    city/state pair is ambiguous. This lets the user drop a marker
    precisely on the site; subsequent GETs use these coords and re-query
    Overpass from the new center.
    """
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    deal.lat = float(data.lat)
    deal.lng = float(data.lng)
    # Invalidate any cached categories - they were centered on the old point.
    ld = deal.location_data or {}
    if isinstance(ld, dict):
        ld.pop("categories", None)
        ld.pop("fetched_at", None)
        deal.location_data = ld
    await db.commit()
    return {"message": "Location updated", "lat": deal.lat, "lng": deal.lng}


# ===== Market Research =====

@router.post("/{deal_id}/market-research", dependencies=[Depends(limit("ai"))])
async def market_research(deal_id: int, db: AsyncSession = Depends(get_db)):
    """Fetch real market data for the deal's city/state via Brave Search + Claude AI."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    city = deal.city or (deal.metrics or {}).get("market_location", {}).get("city", "")
    state = deal.state or (deal.metrics or {}).get("market_location", {}).get("state", "")

    if not city or not state:
        raise HTTPException(status_code=400, detail="City and state are required. Update the deal first.")

    try:
        market_data = await fetch_market_data(city, state)
    except Exception as e:
        msg = str(e)
        status = 503 if ("ANTHROPIC_API_KEY" in msg or "BRAVE_API_KEY" in msg) else 500
        raise HTTPException(status_code=status, detail=f"Market research failed: {msg}")

    # Save to deal metrics
    metrics = deal.metrics.copy() if deal.metrics else {}
    metrics["market_research"] = market_data
    deal.metrics = metrics
    await db.commit()

    return {"message": "Market research complete", "market_research": market_data}


# ===== Cash Flow Projection =====

@router.get("/{deal_id}/cashflow")
async def cashflow_projection(deal_id: int, investment: Optional[float] = None, db: AsyncSession = Depends(get_db)):
    """Generate year-by-year cash flow projections."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if not deal.metrics:
        raise HTTPException(status_code=400, detail="No metrics extracted yet.")

    cashflow = project_cash_flows(deal.metrics, investment_amount=investment)
    return cashflow


# ===== Waterfall Calculator =====

@router.get("/{deal_id}/waterfall")
async def waterfall_calculation(deal_id: int, investment: Optional[float] = None, db: AsyncSession = Depends(get_db)):
    """Calculate waterfall distribution from deal metrics."""
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if not deal.metrics:
        raise HTTPException(status_code=400, detail="No metrics extracted yet.")

    waterfall = waterfall_from_deal(deal.metrics, investment_amount=investment)
    return waterfall


# ===== Comparison =====

@router.post("/compare")
async def compare_deals(data: CompareRequest, db: AsyncSession = Depends(get_db)):
    """Compare multiple deals side-by-side."""
    if len(data.deal_ids) < 2:
        raise HTTPException(status_code=400, detail="Select at least 2 deals to compare")

    result = await db.execute(
        select(Deal)
        .options(selectinload(Deal.developer))
        .where(Deal.id.in_(data.deal_ids))
    )
    deals = result.scalars().all()

    comparison = []
    for deal in deals:
        comparison.append(_deal_to_dict(deal, deal.developer.name if deal.developer else ""))

    return {"deals": comparison}


@router.post("/compare/export")
async def export_comparison(data: CompareRequest, db: AsyncSession = Depends(get_db)):
    """Export deal comparison as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if len(data.deal_ids) < 2:
        raise HTTPException(status_code=400, detail="Select at least 2 deals to compare")

    result = await db.execute(
        select(Deal)
        .options(selectinload(Deal.developer))
        .where(Deal.id.in_(data.deal_ids))
    )
    deals = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal Comparison"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    section_font = Font(bold=True, size=11, color="4361ee")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Header row
    ws.cell(row=1, column=1, value="Metric").font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].border = thin_border
    ws.column_dimensions["A"].width = 30

    for col_idx, deal in enumerate(deals, 2):
        cell = ws.cell(row=1, column=col_idx, value=deal.project_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    # Define metric rows by section
    sections = [
        ("SCORES", [
            ("Overall Score", lambda d: (d.scores or {}).get("overall")),
            ("Returns Score", lambda d: ((d.scores or {}).get("returns") or {}).get("score")),
            ("Market Score", lambda d: ((d.scores or {}).get("market") or {}).get("score")),
            ("Structure Score", lambda d: ((d.scores or {}).get("structure") or {}).get("score")),
            ("Risk Score", lambda d: ((d.scores or {}).get("risk") or {}).get("score")),
            ("Financials Score", lambda d: ((d.scores or {}).get("financials") or {}).get("score")),
        ]),
        ("DEAL STRUCTURE", [
            ("Investment Class", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("investment_class")),
            ("Minimum Investment", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("minimum_investment")),
            ("Total Project Cost", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("total_project_cost")),
            ("Total Equity", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("total_equity_required")),
            ("Debt Amount", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("debt_amount")),
            ("LTV", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("ltv")),
            ("Interest Rate", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("interest_rate")),
            ("Hold Period (yrs)", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("hold_period_years")),
            ("Preferred Return", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("preferred_return")),
            ("GP Co-Invest", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("gp_coinvest")),
            ("Asset Mgmt Fee", lambda d: ((d.metrics or {}).get("deal_structure") or {}).get("fees_asset_mgmt")),
        ]),
        ("TARGET RETURNS", [
            ("Target IRR", lambda d: canonical_return_summary(d.metrics or {}).get("target_irr")),
            ("Equity Multiple", lambda d: canonical_return_summary(d.metrics or {}).get("target_equity_multiple")),
            ("Cash-on-Cash", lambda d: canonical_return_summary(d.metrics or {}).get("cash_on_cash")),
            ("Avg Annual Return", lambda d: ((d.metrics or {}).get("target_returns") or {}).get("target_avg_annual_return")),
            ("Projected Profit", lambda d: ((d.metrics or {}).get("target_returns") or {}).get("projected_profit")),
        ]),
        ("PROJECT DETAILS", [
            ("Unit Count", lambda d: ((d.metrics or {}).get("project_details") or {}).get("unit_count")),
            ("Total SqFt", lambda d: ((d.metrics or {}).get("project_details") or {}).get("total_sqft")),
            ("Price/Unit", lambda d: ((d.metrics or {}).get("project_details") or {}).get("price_per_unit")),
            ("Price/SqFt", lambda d: ((d.metrics or {}).get("project_details") or {}).get("price_per_sqft")),
            ("Construction Type", lambda d: ((d.metrics or {}).get("project_details") or {}).get("construction_type")),
            ("Entitlement Status", lambda d: ((d.metrics or {}).get("project_details") or {}).get("entitlement_status")),
        ]),
        ("CONSTRUCTION COSTS", [
            ("Total Project Cost", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("total_project_cost")),
            ("Total Cost/Unit", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("total_project_cost_per_unit")),
            ("Hard Costs Total", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("hard_costs_total")),
            ("Hard Costs/Unit", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("hard_costs_per_unit")),
            ("Hard Costs/SqFt", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("hard_costs_per_sqft")),
            ("Land Cost", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("land_cost_total")),
            ("Land Cost/Unit", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("land_cost_per_unit")),
            ("Soft Costs", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("soft_costs_total")),
            ("Contingency", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("contingency_total")),
            ("Contingency %", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("contingency_pct")),
            ("Financing Costs", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("financing_costs_total")),
            ("Developer Fee", lambda d: ((d.metrics or {}).get("construction_costs") or {}).get("developer_fee_total")),
        ]),
        ("FINANCIAL PROJECTIONS", [
            ("Stabilized NOI", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("stabilized_noi")),
            ("Entry Cap Rate", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("entry_cap_rate")),
            ("Exit Cap Rate", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("exit_cap_rate")),
            ("Avg Rent/Unit", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("avg_rent_per_unit")),
            ("Rent Growth", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("rent_growth_assumption")),
            ("Occupancy", lambda d: ((d.metrics or {}).get("financial_projections") or {}).get("occupancy_assumption")),
        ]),
    ]

    row = 2
    green_fill = PatternFill(start_color="ecfdf5", end_color="ecfdf5", fill_type="solid")
    red_fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")

    for section_name, metrics_list in sections:
        # Section header
        cell = ws.cell(row=row, column=1, value=section_name)
        cell.font = section_font
        row += 1

        for metric_name, getter in metrics_list:
            ws.cell(row=row, column=1, value=metric_name).border = thin_border
            values = []
            for col_idx, deal in enumerate(deals, 2):
                val = getter(deal)
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if isinstance(val, (int, float)):
                    values.append((col_idx, val))

            # Highlight best/worst if numeric
            if len(values) >= 2:
                best_col = max(values, key=lambda x: x[1])[0]
                worst_col = min(values, key=lambda x: x[1])[0]
                if best_col != worst_col:
                    ws.cell(row=row, column=best_col).fill = green_fill
                    ws.cell(row=row, column=worst_col).fill = red_fill
            row += 1
        row += 1  # Space between sections

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deal_comparison.xlsx"},
    )

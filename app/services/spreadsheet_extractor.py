"""Spreadsheet extraction service for deal underwriting files."""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Any


MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 180
MAX_COLS_PER_SHEET = 50
MAX_CELLS_FOR_PROVENANCE = 1500
MAX_KEY_ROWS = 250

KEY_TERMS = {
    "irr",
    "return",
    "multiple",
    "equity",
    "cash",
    "yield",
    "distribution",
    "pref",
    "preferred",
    "debt",
    "loan",
    "ltv",
    "dscr",
    "noi",
    "cap",
    "cost",
    "budget",
    "hard",
    "soft",
    "land",
    "contingency",
    "rent",
    "occupancy",
    "units",
    "sponsor",
    "gp",
    "fee",
    "waterfall",
    "hold",
    "sale",
    "exit",
}


@dataclass
class SpreadsheetExtractionResult:
    text: str = ""
    page_count: int = 0
    ocr_page_count: int = 0
    tables: list[dict] = field(default_factory=list)
    images: list[dict] = field(default_factory=list)
    cells: list[dict] = field(default_factory=list)
    key_rows: list[dict] = field(default_factory=list)
    page_diagnostics: list[dict] = field(default_factory=list)
    quality_score: int = 100


def extract_spreadsheet(file_path: str) -> SpreadsheetExtractionResult:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in {".xlsx", ".xlsm"}:
        return _extract_xlsx(file_path)
    if ext == ".xls":
        return _extract_xls(file_path)
    if ext == ".csv":
        return _extract_csv(file_path)

    result = SpreadsheetExtractionResult()
    result.text = f"[Unsupported spreadsheet type: {ext or 'unknown'}]"
    result.quality_score = 0
    return result


def _extract_xlsx(file_path: str) -> SpreadsheetExtractionResult:
    from openpyxl import load_workbook

    result = SpreadsheetExtractionResult()
    wb = load_workbook(file_path, data_only=True, read_only=True)
    formula_wb = None
    try:
        formula_wb = load_workbook(file_path, data_only=False, read_only=True)
    except Exception:
        formula_wb = None

    sheet_names = list(wb.sheetnames)[:MAX_SHEETS]
    text_parts: list[str] = []
    try:
        for sheet_index, sheet_name in enumerate(sheet_names, 1):
            ws = wb[sheet_name]
            formula_ws = formula_wb[sheet_name] if formula_wb and sheet_name in formula_wb.sheetnames else None
            rows, table_rows, cells, key_rows = _worksheet_rows(ws, formula_ws, sheet_name)
            result.cells.extend(cells)
            result.key_rows.extend(key_rows)
            result.page_diagnostics.append(
                {"page": sheet_index, "source": "spreadsheet", "chars": sum(len(r) for r in rows)}
            )
            if table_rows:
                result.tables.append({"sheet": sheet_name, "rows": table_rows})
            text_parts.append(_render_sheet(sheet_name, rows))
    finally:
        wb.close()
        if formula_wb:
            formula_wb.close()

    result.cells = result.cells[:MAX_CELLS_FOR_PROVENANCE]
    result.key_rows = result.key_rows[:MAX_KEY_ROWS]
    result.page_count = len(sheet_names)
    body = "\n\n".join(part for part in text_parts if part.strip())
    result.text = _render_key_rows(result.key_rows) + ("\n\n" if body else "") + body
    result.quality_score = 100 if result.text.strip() else 0
    return result


def _extract_xls(file_path: str) -> SpreadsheetExtractionResult:
    result = SpreadsheetExtractionResult()
    try:
        import xlrd
    except ImportError:
        result.text = "[Legacy .xls support requires xlrd.]"
        result.quality_score = 0
        return result

    book = xlrd.open_workbook(file_path)
    text_parts: list[str] = []
    sheet_count = min(book.nsheets, MAX_SHEETS)
    for sheet_index in range(sheet_count):
        sheet = book.sheet_by_index(sheet_index)
        rows: list[str] = []
        table_rows: list[list[Any]] = []
        max_rows = min(sheet.nrows, MAX_ROWS_PER_SHEET)
        max_cols = min(sheet.ncols, MAX_COLS_PER_SHEET)
        for row_idx in range(max_rows):
            values = [_clean_cell(sheet.cell_value(row_idx, col_idx)) for col_idx in range(max_cols)]
            rendered_values = [_render_cell(v) for v in values]
            if not any(v != "" for v in rendered_values):
                continue
            table_rows.append(values)
            rows.append("\t".join(str(v) for v in rendered_values))
            if _is_key_row(rendered_values):
                result.key_rows.append(
                    {"sheet": sheet.name, "row": row_idx + 1, "values": rendered_values, "cells": []}
                )
            for col_idx, value in enumerate(values, 1):
                if value in (None, "") or len(result.cells) >= MAX_CELLS_FOR_PROVENANCE:
                    continue
                result.cells.append(
                    {"sheet": sheet.name, "cell": _cell_ref(row_idx + 1, col_idx), "value": value}
                )
        if table_rows:
            result.tables.append({"sheet": sheet.name, "rows": table_rows})
        result.page_diagnostics.append(
            {"page": sheet_index + 1, "source": "spreadsheet", "chars": sum(len(r) for r in rows)}
        )
        text_parts.append(_render_sheet(sheet.name, rows))

    result.key_rows = result.key_rows[:MAX_KEY_ROWS]
    result.page_count = sheet_count
    body = "\n\n".join(part for part in text_parts if part.strip())
    result.text = _render_key_rows(result.key_rows) + ("\n\n" if body else "") + body
    result.quality_score = 100 if result.text.strip() else 0
    return result


def _extract_csv(file_path: str) -> SpreadsheetExtractionResult:
    result = SpreadsheetExtractionResult(page_count=1)
    rows: list[str] = []
    table_rows: list[list[str]] = []
    with open(file_path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.reader(fh)
        for idx, row in enumerate(reader):
            if idx >= MAX_ROWS_PER_SHEET:
                break
            trimmed = row[:MAX_COLS_PER_SHEET]
            rendered = [_render_cell(cell) for cell in trimmed]
            if not any(cell.strip() for cell in rendered):
                continue
            table_rows.append(trimmed)
            rows.append("\t".join(rendered))
            if _is_key_row(rendered):
                result.key_rows.append({"sheet": "CSV", "row": idx + 1, "values": rendered, "cells": []})
            for col_idx, value in enumerate(trimmed, 1):
                if not str(value).strip() or len(result.cells) >= MAX_CELLS_FOR_PROVENANCE:
                    continue
                result.cells.append(
                    {"sheet": "CSV", "cell": _cell_ref(idx + 1, col_idx), "value": value}
                )
    if table_rows:
        result.tables.append({"sheet": "CSV", "rows": table_rows})
    result.key_rows = result.key_rows[:MAX_KEY_ROWS]
    result.page_diagnostics.append({"page": 1, "source": "spreadsheet", "chars": sum(len(r) for r in rows)})
    body = _render_sheet(os.path.basename(file_path), rows)
    result.text = _render_key_rows(result.key_rows) + ("\n\n" if body else "") + body
    result.quality_score = 100 if result.text.strip() else 0
    return result


def _worksheet_rows(ws, formula_ws, sheet_name: str) -> tuple[list[str], list[list[Any]], list[dict], list[dict]]:
    rows: list[str] = []
    table_rows: list[list[Any]] = []
    cells: list[dict] = []
    key_rows: list[dict] = []
    max_row = min(ws.max_row or 0, MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 0, MAX_COLS_PER_SHEET)

    for row_idx in range(1, max_row + 1):
        row_values: list[str] = []
        raw_values: list[Any] = []
        row_cells: list[dict] = []
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            formula = formula_ws.cell(row=row_idx, column=col_idx).value if formula_ws else None
            rendered = _render_cell(value)
            row_values.append(rendered)
            raw_values.append(value)
            if rendered != "":
                cell = {
                    "sheet": sheet_name,
                    "cell": _cell_ref(row_idx, col_idx),
                    "value": value,
                }
                if isinstance(formula, str) and formula.startswith("="):
                    cell["formula"] = formula
                row_cells.append(cell)
                if len(cells) < MAX_CELLS_FOR_PROVENANCE:
                    cells.append(cell)
        if not any(v != "" for v in row_values):
            continue
        table_rows.append(raw_values)
        rows.append("\t".join(row_values))
        if _is_key_row(row_values):
            key_rows.append(
                {"sheet": sheet_name, "row": row_idx, "values": row_values, "cells": row_cells[:20]}
            )

    return rows, table_rows, cells, key_rows


def _render_sheet(sheet_name: str, rows: list[str]) -> str:
    if not rows:
        return f"--- Sheet: {sheet_name} ---\n[No readable cells found]"
    return f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows)


def _render_key_rows(key_rows: list[dict]) -> str:
    if not key_rows:
        return ""
    lines = [
        "--- KEY SPREADSHEET ROWS ---",
        "These rows contain financial terms or formulas and should be checked first before reading full sheets.",
    ]
    for item in key_rows[:MAX_KEY_ROWS]:
        sheet = item.get("sheet") or "Sheet"
        row = item.get("row") or "?"
        values = [str(v) for v in (item.get("values") or []) if str(v).strip()]
        cells = item.get("cells") or []
        cell_refs = ", ".join(
            f"{c.get('cell')}={_render_cell(c.get('value'))}"
            + (f" ({c.get('formula')})" if c.get("formula") else "")
            for c in cells[:10]
        )
        rendered_values = " | ".join(values[:20])
        suffix = f" [{cell_refs}]" if cell_refs else ""
        lines.append(f"{sheet}!row {row}: {rendered_values}{suffix}")
    return "\n".join(lines)


def _is_key_row(values: list[str]) -> bool:
    text = " ".join(str(v).lower() for v in values if v)
    if not text:
        return False
    return any(term in text for term in KEY_TERMS)


def _cell_ref(row_idx: int, col_idx: int) -> str:
    try:
        from openpyxl.utils import get_column_letter

        col = get_column_letter(col_idx)
    except Exception:
        col = str(col_idx)
    return f"{col}{row_idx}"


def _render_cell(value: Any) -> str:
    return str(_clean_cell(value))


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 6)
    return value
